# db_to_xlsx.py
"""
Convert reports_binder.db (SQLite) -> formatted Excel (.xlsx)

Usage:
    python db_to_xlsx.py /path/to/reports_binder.db /path/to/output.xlsx

What you get:
- Sheet "Reports": ordered columns, filters, frozen header, zebra stripes, date formatting
- Sheet "Failures" (if present): same niceties
- Sheet "Progress" (if present): same niceties
- Sheet "Summary": quick high-level pivots on Reports
"""

import sys
import sqlite3
from pathlib import Path
from typing import List, Tuple

import pandas as pd

# Choose engine: openpyxl (installed by default in many envs) for styling
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------- Helpers ----------
def read_table_names(conn: sqlite3.Connection) -> List[str]:
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")
    return [r[0] for r in cur.fetchall()]


def read_df(conn: sqlite3.Connection, table: str) -> pd.DataFrame:
    return pd.read_sql_query(f"SELECT * FROM {table};", conn)


def to_datetime_safe(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce", utc=False).dt.date
    return df


def order_reports_columns(df: pd.DataFrame) -> pd.DataFrame:
    preferred = [
        "pdf_file", "page_label",
        "date", "date_sent",
        "work_order", "engineer_initials", "billing",
        "id",
    ]
    existing = [c for c in preferred if c in df.columns]
    rest = [c for c in df.columns if c not in existing]
    return df[existing + rest]


def auto_widths(ws, min_width=10, max_width=50):
    # Estimate width based on max string length in each column (including header)
    for col_idx, col_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        header = col_cells[0].value or ""
        max_len = len(str(header))
        for cell in col_cells[1:]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_header_row(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")  # deep blue
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center


def add_table(ws, table_name: str):
    # Create an Excel "Table" for filters and banded rows
    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row
    ref = f"A1:{last_col}{last_row}"
    tbl = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)


def freeze_panes(ws):
    ws.freeze_panes = "A2"  # freeze top row


def thin_border_range(ws):
    thin = Side(style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


def format_date_columns(ws, header_row=1):
    # Detect columns whose header contains 'date' (case-insensitive) and apply date format
    headers = {cell.value: idx for idx, cell in enumerate(ws[header_row], start=1)}
    for name, idx in headers.items():
        if not isinstance(name, str):
            continue
        if "date" in name.lower():
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = "yyyy-mm-dd"


def write_sheet(writer, name: str, df: pd.DataFrame):
    df.to_excel(writer, index=False, sheet_name=name)
    ws = writer.book[name]

    style_header_row(ws)
    add_table(ws, table_name=f"{name.replace(' ', '')}Table")
    freeze_panes(ws)
    thin_border_range(ws)
    auto_widths(ws)
    format_date_columns(ws)


def make_summary(df_reports: pd.DataFrame) -> pd.DataFrame:
    # Simple, helpful pivots: total rows by pdf_file, and by year
    out = []

    if "pdf_file" in df_reports.columns:
        by_pdf = df_reports.groupby("pdf_file", dropna=False).size().reset_index(name="rows")
        by_pdf = by_pdf.sort_values("rows", ascending=False)
        out.append(("Rows by PDF", by_pdf))

    if "date" in df_reports.columns:
        tmp = df_reports.copy()
        tmp["year"] = pd.to_datetime(tmp["date"], errors="coerce").dt.year
        by_year = tmp.groupby("year", dropna=False).size().reset_index(name="rows")
        by_year = by_year.sort_values("year", na_position="last")
        out.append(("Rows by Year (date)", by_year))

    if "work_order" in df_reports.columns:
        wo_top = (
            df_reports.groupby("work_order", dropna=False)
            .size()
            .reset_index(name="rows")
            .sort_values("rows", ascending=False)
            .head(50)
        )
        out.append(("Top Work Orders (by rows)", wo_top))

    # Stitch into a single sheet with section titles
    if not out:
        return pd.DataFrame()

    # Create a flattened DataFrame with section separators
    blocks = []
    for title, block in out:
        sep = pd.DataFrame({block.columns[0]: [f"— {title} —"]})
        blocks += [sep, block, pd.DataFrame([[]])]  # blank line
    return pd.concat(blocks, ignore_index=True, sort=False)


# --- Add/replace the start of main() with this ---
def main():
    # Defaults if no args provided
    default_db = Path("../uploads/reports_binder.db").resolve()
    default_xlsx = Path("../uploads/reports_binder.xlsx").resolve()

    if len(sys.argv) == 1:
        db_path = default_db
        xlsx_path = default_xlsx
        print(f"No args provided. Using defaults:\n  DB:   {db_path}\n  XLSX: {xlsx_path}")
    elif len(sys.argv) == 2:
        db_path = Path(sys.argv[1]).resolve()
        # Put XLSX next to the DB with same stem
        xlsx_path = db_path.with_suffix(".xlsx")
        print(f"One arg provided. Using:\n  DB:   {db_path}\n  XLSX: {xlsx_path}")
    else:
        db_path = Path(sys.argv[1]).resolve()
        xlsx_path = Path(sys.argv[2]).resolve()

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if not db_path.exists():
        print(f"ERROR: DB not found at {db_path}")
        sys.exit(2)

    with sqlite3.connect(str(db_path)) as conn:
        tables = read_table_names(conn)
        reports = read_df(conn, "reports") if "reports" in tables else pd.DataFrame()
        failures = read_df(conn, "failures") if "failures" in tables else pd.DataFrame()
        progress = read_df(conn, "progress") if "progress" in tables else pd.DataFrame()

    if not reports.empty:
        reports = to_datetime_safe(reports, ["date", "date_sent"])
        reports = order_reports_columns(reports)

    summary = make_summary(reports) if not reports.empty else pd.DataFrame({"Info": ["No data in reports"]})

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        if not reports.empty:
            write_sheet(writer, "Reports", reports)
        else:
            pd.DataFrame({"Info": ["No reports found"]}).to_excel(writer, index=False, sheet_name="Reports")

        if not failures.empty:
            write_sheet(writer, "Failures", failures)
        if not progress.empty:
            write_sheet(writer, "Progress", progress)

        if not summary.empty:
            write_sheet(writer, "Summary", summary)
        else:
            pd.DataFrame({"Info": ["No summary available"]}).to_excel(writer, index=False, sheet_name="Summary")

    print(f"✅ Wrote Excel to: {xlsx_path}")



if __name__ == "__main__":
    main()
